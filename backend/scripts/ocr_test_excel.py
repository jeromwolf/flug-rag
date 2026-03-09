"""Generate Excel report for Upstage OCR test results."""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

with open("data/upstage_ocr_test_results.json", "r") as f:
    results = json.load(f)

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2F5496")
title_font = Font(bold=True, size=14)
subtitle_font = Font(bold=True, size=11, color="2F5496")
border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
success_fill = PatternFill("solid", fgColor="C6EFCE")
fail_fill = PatternFill("solid", fgColor="FFC7CE")
warn_fill = PatternFill("solid", fgColor="FFEB9C")
num_fmt = "#,##0"

# File paths for each test
file_paths = [
    "data/uploads/인쇄홍보물/인쇄홍보물/[한국가스기술공사] 브로슈어(국문)_2024.pdf",
    "data/sample_dataset/한국가스공사법/한국가스공사법(법률)(제13160호)(20150203).pdf",
    "data/uploads/ALIO_한국가스기술공사_검색결과/ALIO_한국가스기술공사_검색결과/다운로드_전체통합/2023년 반기 재무상태표.pdf",
    "data/uploads/ALIO_한국가스기술공사_검색결과/ALIO_한국가스기술공사_검색결과/다운로드_전체통합/2021년 감사보고서.pdf",
    "data/uploads/ALIO_한국가스기술공사_검색결과/ALIO_한국가스기술공사_검색결과/다운로드_전체통합/2025년도 제334회 이사회 회의록.pdf",
    "data/uploads/국외출장_결과보고서/국외출장_결과보고서/50까지/요약출장보고서_해외기술자문용역 중간회의 출장(일본_ 오사카).pdf",
    "data/uploads/ALIO_한국가스기술공사_검색결과/ALIO_한국가스기술공사_검색결과/다운로드_전체통합/2023년도 인권경영보고서.pdf",
    "data/uploads/ALIO_한국가스기술공사_검색결과/ALIO_한국가스기술공사_검색결과/다운로드_전체통합/2025년도 일상감사 준수실태 특정감사 결과 .pdf",
    "data/uploads/ALIO_한국가스기술공사_검색결과/ALIO_한국가스기술공사_검색결과/다운로드_전체통합/한국가스공사법 시행령 (1).pdf",
    "data/uploads/국외출장_결과보고서/국외출장_결과보고서/50까지/국외출장보고_베트남 YHP Refrigerated LPG 저장탱크 설계용역 Tank Foundation 및 Pile Deviation 대면회의 참석 관련 국외출장 결과보고_게시용.pdf",
]

# Verification questions & answers for each test
qa_data = [
    {
        "question": "브로셔에서 한국가스기술공사의 본사 주소와 전화번호가 정확히 추출되었는가?",
        "answer": "△ — 주소/TEL/FAX 텍스트는 존재하나 영문 혼재로 구조 깨짐. 'TEL.' 뒤에 웹사이트(kogas-tech.or.kr)가 오고 번호(042.600.8000/8195)는 'FAX.' 뒤에 뒤섞임. 이미지+텍스트 혼합 문서의 OCR 한계",
        "pass": False,
    },
    {
        "question": "한국가스공사법의 법률 번호와 시행일, 제1조 목적 조항이 정확히 추출되었는가?",
        "answer": "O — '[법률 제13160호, 2015. 2. 3., 일부개정]', 제1조~제10조 목차 정확, 조문 텍스트 완전 추출",
        "pass": True,
    },
    {
        "question": "재무상태표의 회계 기간, 회사명, 단위(원), 표 구조(과목/주석/당기/전기)가 정확한가?",
        "answer": "O — '제31(당)기 반기말 2023년 6월 30일', '주식회사 한국가스기술공사 (단위:원)', 자산/부채 표 구조 HTML 정확 추출",
        "pass": True,
    },
    {
        "question": "감사보고서에서 감사법인명, 대상 기간(제29기), 감사의견이 추출되었는가?",
        "answer": "O — '우리회계법인', '제29(당)기 2021년 01월 01일~2021년 12월 31일', 89페이지 105,252자 완전 추출, 표 포함",
        "pass": True,
    },
    {
        "question": "이사회 회의록에서 회의일시, 장소, 출석임원 명단이 정확히 추출되었는가?",
        "answer": "O — '2025년 10월 27일(월), 10:50~12:30', '본사 4층 영상회의실', 상임이사(진수남, 홍광희), 비상임이사(왕제필 외 3인) 정확",
        "pass": True,
    },
    {
        "question": "출장보고서(일본)에서 출장 목적, 일정, 출장지역이 정확히 추출되었는가?",
        "answer": "O — '동북아LNG터미널 설계감리 해외기술자문 용역', '2025.5.26(월)~5.28(수) [2박3일]', '일본 (오사카)' 정확",
        "pass": True,
    },
    {
        "question": "인권경영보고서에서 보고서 제목과 연도가 추출되었는가? 디자인 요소 노이즈가 있는가?",
        "answer": "△ — '2023년도 인권경영보고서' 제목 추출되었으나, 표지 디자인 텍스트 반복 노이즈 있음 (제목 10회 반복)",
        "pass": False,
    },
    {
        "question": "특정감사 결과에서 감사목적, 감사범위, 감사기간, 감사결과 표가 정확한가?",
        "answer": "O — '일상감사 제도 효과성 제고', '전 부서', '2025.10.27~10.31', 감사결과 표(구분/건수/세부내용) HTML 추출",
        "pass": True,
    },
    {
        "question": "시행령에서 대통령령 번호, 시행일, 소관부서, 제1조 목적이 정확히 추출되었는가?",
        "answer": "O — '[대통령령 제31176호, 2020.11.24., 타법개정]', '산업통상자원부(가스산업과) 044-203-5231', 제1조 정확",
        "pass": True,
    },
    {
        "question": "출장보고서(베트남)에서 출장 목적, 일정, 출장지역, 기술 용어(Tank Foundation 등)가 정확한가?",
        "answer": "O — 'Tank Foundation 및 Pile Deviation 대면회의', '24.04.08(월)~04.12(금) [4박5일]', 'YHP Refrigerated LPG' 기술용어 정확",
        "pass": True,
    },
]

total_pages = sum(r.get("pages", 0) for r in results)
total_text = sum(r.get("text_len", 0) for r in results)
total_time = sum(r.get("elapsed", 0) for r in results)
tables_count = sum(1 for r in results if r.get("has_tables"))
pass_count = sum(1 for q in qa_data if q["pass"])


# ══════════════════════════════════════════════════════════════
# Sheet 1: 테스트 요약
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "테스트 요약"

ws1.merge_cells("A1:N1")
ws1["A1"] = "업스테이지 Document Parse API OCR 테스트 리포트"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(horizontal="center")

ws1.merge_cells("A2:N2")
ws1["A2"] = "테스트일: 2026-03-02 | API: document-parse | 총 10건 | 실제 비용: $1.84 (확인완료)"
ws1["A2"].font = Font(size=10, color="666666")
ws1["A2"].alignment = Alignment(horizontal="center")

# Summary row
ws1["A4"] = "성공률"
ws1["A4"].font = subtitle_font
ws1["B4"] = "10/10 (100%)"
ws1["B4"].font = Font(bold=True, size=12, color="006100")
ws1["B4"].fill = success_fill

ws1["C4"] = "품질 검증"
ws1["C4"].font = subtitle_font
ws1["D4"] = f"{pass_count}/10 통과"
ws1["D4"].font = Font(bold=True, size=12, color="006100" if pass_count >= 9 else "C00000")

ws1["E4"] = "총 페이지"
ws1["E4"].font = subtitle_font
ws1["F4"] = f"{total_pages}p"
ws1["F4"].font = Font(bold=True, size=12)

ws1["G4"] = "실제 비용"
ws1["G4"].font = subtitle_font
ws1["H4"] = "$1.84 (확인)"
ws1["H4"].font = Font(bold=True, size=12, color="C00000")

ws1["I4"] = "페이지당 단가"
ws1["I4"].font = subtitle_font
ws1["J4"] = "$0.01/page"
ws1["J4"].font = Font(bold=True, size=12)

# Headers (row 6)
headers = [
    "#", "문서 유형", "파일명", "파일 경로",
    "크기(KB)", "페이지", "텍스트(chars)", "HTML(chars)",
    "표 인식", "소요시간(초)", "상태",
    "검증 질문", "검증 결과", "통과",
]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=6, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = border

# Data rows
for i, r in enumerate(results):
    row = 7 + i
    qa = qa_data[i]

    ws1.cell(row=row, column=1, value=r["test"]).border = border
    ws1.cell(row=row, column=1).alignment = Alignment(horizontal="center")

    ws1.cell(row=row, column=2, value=r["desc"]).border = border

    ws1.cell(row=row, column=3, value=r.get("file", "")).border = border

    c = ws1.cell(row=row, column=4, value=file_paths[i])
    c.border = border
    c.font = Font(size=9, color="0563C1", underline="single")

    c = ws1.cell(row=row, column=5, value=r.get("size_kb", 0))
    c.number_format = num_fmt
    c.border = border
    c.alignment = Alignment(horizontal="right")

    c = ws1.cell(row=row, column=6, value=r.get("pages", 0))
    c.border = border
    c.alignment = Alignment(horizontal="center")

    c = ws1.cell(row=row, column=7, value=r.get("text_len", 0))
    c.number_format = num_fmt
    c.border = border
    c.alignment = Alignment(horizontal="right")

    c = ws1.cell(row=row, column=8, value=r.get("html_len", 0))
    c.number_format = num_fmt
    c.border = border
    c.alignment = Alignment(horizontal="right")

    has_table = r.get("has_tables", False)
    c = ws1.cell(row=row, column=9, value="O" if has_table else "X")
    c.border = border
    c.alignment = Alignment(horizontal="center")
    c.font = Font(bold=True, color="006100" if has_table else "C00000")

    c = ws1.cell(row=row, column=10, value=r.get("elapsed", 0))
    c.number_format = "0.0"
    c.border = border
    c.alignment = Alignment(horizontal="center")

    status = r.get("status", "")
    c = ws1.cell(row=row, column=11, value=status)
    c.border = border
    c.fill = success_fill if status == "SUCCESS" else warn_fill
    c.alignment = Alignment(horizontal="center")

    # Verification Q&A
    c = ws1.cell(row=row, column=12, value=qa["question"])
    c.border = border
    c.alignment = Alignment(wrap_text=True, vertical="top")

    c = ws1.cell(row=row, column=13, value=qa["answer"])
    c.border = border
    c.alignment = Alignment(wrap_text=True, vertical="top")

    c = ws1.cell(row=row, column=14, value="PASS" if qa["pass"] else "WARN")
    c.border = border
    c.alignment = Alignment(horizontal="center")
    c.font = Font(bold=True, color="006100" if qa["pass"] else "BF8F00")
    c.fill = success_fill if qa["pass"] else warn_fill

    ws1.row_dimensions[row].height = 50

# Totals row
tr = 17
for col in range(1, 15):
    ws1.cell(row=tr, column=col, value="").border = border
ws1.cell(row=tr, column=1, value="합계").font = Font(bold=True)
ws1.cell(row=tr, column=5, value=sum(r.get("size_kb", 0) for r in results)).font = Font(bold=True)
ws1.cell(row=tr, column=5).number_format = num_fmt
ws1.cell(row=tr, column=6, value=total_pages).font = Font(bold=True)
ws1.cell(row=tr, column=6).alignment = Alignment(horizontal="center")
ws1.cell(row=tr, column=7, value=total_text).font = Font(bold=True)
ws1.cell(row=tr, column=7).number_format = num_fmt
ws1.cell(row=tr, column=8, value=sum(r.get("html_len", 0) for r in results)).font = Font(bold=True)
ws1.cell(row=tr, column=8).number_format = num_fmt
ws1.cell(row=tr, column=9, value=f"{tables_count}/10").font = Font(bold=True)
ws1.cell(row=tr, column=10, value=round(total_time, 1)).font = Font(bold=True)
ws1.cell(row=tr, column=14, value=f"{pass_count}/10").font = Font(bold=True, size=12)

# Column widths
widths = [4, 30, 40, 55, 10, 8, 12, 12, 8, 10, 10, 50, 60, 8]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════
# Sheet 2: OCR 텍스트 미리보기
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("OCR 텍스트 미리보기")

headers2 = ["#", "문서 유형", "파일명", "파일 경로", "OCR 추출 텍스트 (첫 500자)"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

for i, r in enumerate(results):
    row = 2 + i
    ws2.cell(row=row, column=1, value=r["test"]).border = border
    ws2.cell(row=row, column=2, value=r["desc"]).border = border
    ws2.cell(row=row, column=3, value=r.get("file", "")).border = border
    c = ws2.cell(row=row, column=4, value=file_paths[i])
    c.border = border
    c.font = Font(size=9, color="0563C1", underline="single")
    c = ws2.cell(row=row, column=5, value=r.get("preview", "")[:500])
    c.border = border
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[row].height = 80

ws2.column_dimensions["A"].width = 4
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 40
ws2.column_dimensions["D"].width = 55
ws2.column_dimensions["E"].width = 120


# ══════════════════════════════════════════════════════════════
# Sheet 3: 비용 분석
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("비용 분석")

ws3.merge_cells("A1:D1")
ws3["A1"] = "업스테이지 API 비용 분석"
ws3["A1"].font = title_font

# Confirmed balance
ws3["A3"] = "크레딧 현황 (실제 확인)"
ws3["A3"].font = subtitle_font

bal_headers = ["항목", "금액", "비고"]
for col, h in enumerate(bal_headers, 1):
    cell = ws3.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

balance_items = [
    ("총 크레딧", 10.00, "2026-03-01 UTC 충전"),
    ("테스트 전 잔액", 7.24, "스크린샷 확인"),
    ("이번 테스트 사용액", 1.84, "184페이지 x $0.01/page"),
    ("테스트 후 잔액", 5.40, "스크린샷 확인 (정확히 일치)"),
    ("페이지당 단가 (확인)", 0.01, "$1.84 / 184p = $0.01/page"),
]
for i, (label, val, note) in enumerate(balance_items):
    row = 5 + i
    ws3.cell(row=row, column=1, value=label).border = border
    c = ws3.cell(row=row, column=2, value=val)
    c.number_format = "$#,##0.00"
    c.border = border
    if i == 3:
        c.font = Font(bold=True, size=12, color="2F5496")
    ws3.cell(row=row, column=3, value=note).border = border
    ws3.cell(row=row, column=3).font = Font(size=9, color="666666")

# Test cost detail
ws3["A11"] = "건별 비용 상세"
ws3["A11"].font = subtitle_font

cost_headers = ["#", "문서 유형", "파일명", "페이지", "비용"]
for col, h in enumerate(cost_headers, 1):
    cell = ws3.cell(row=12, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border

for i, r in enumerate(results):
    row = 13 + i
    ws3.cell(row=row, column=1, value=r["test"]).border = border
    ws3.cell(row=row, column=2, value=r["desc"]).border = border
    ws3.cell(row=row, column=3, value=r.get("file", "")).border = border
    pages = r.get("pages", 0)
    ws3.cell(row=row, column=4, value=pages).border = border
    ws3.cell(row=row, column=4).alignment = Alignment(horizontal="center")
    c = ws3.cell(row=row, column=5, value=pages * 0.01)
    c.number_format = "$#,##0.00"
    c.border = border

ctr = 23
for col in range(1, 6):
    ws3.cell(row=ctr, column=col, value="").border = border
ws3.cell(row=ctr, column=2, value="합계").font = Font(bold=True)
ws3.cell(row=ctr, column=4, value=total_pages).font = Font(bold=True)
ws3.cell(row=ctr, column=4).alignment = Alignment(horizontal="center")
c = ws3.cell(row=ctr, column=5, value=total_pages * 0.01)
c.number_format = "$#,##0.00"
c.font = Font(bold=True, size=12, color="C00000")

# Projection
ws3["A25"] = "발주처 데이터 전체 OCR 비용 예측"
ws3["A25"].font = subtitle_font

proj_headers = ["데이터셋", "파일 수", "예상 페이지", "예상 비용", "비고"]
for col, h in enumerate(proj_headers, 1):
    cell = ws3.cell(row=26, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.border = border

datasets = [
    ("내부규정 (HWP)", 676, 3384, 33.84, "평균 5p/file, HWP 지원 여부 미팅 확인"),
    ("인쇄홍보물 (PDF)", 7, 154, 1.54, "평균 22p/file (브로셔 기준)"),
    ("ALIO 공시 (PDF/XLSX)", 403, 1209, 12.09, "평균 3p/file"),
    ("국외출장 보고서 (PDF)", 100, 500, 5.00, "평균 5p/file, 스캔 PDF 포함"),
    ("한국가스공사법 (PDF)", 5, 30, 0.30, "평균 6p/file"),
]
for i, (name, files, pages, cost, note) in enumerate(datasets):
    row = 27 + i
    ws3.cell(row=row, column=1, value=name).border = border
    ws3.cell(row=row, column=2, value=files).border = border
    ws3.cell(row=row, column=2).number_format = num_fmt
    ws3.cell(row=row, column=3, value=pages).border = border
    ws3.cell(row=row, column=3).number_format = num_fmt
    c = ws3.cell(row=row, column=4, value=cost)
    c.number_format = "$#,##0.00"
    c.border = border
    ws3.cell(row=row, column=5, value=note).border = border
    ws3.cell(row=row, column=5).font = Font(size=9, color="666666")

ptr = 32
for col in range(1, 6):
    ws3.cell(row=ptr, column=col, value="").border = border
ws3.cell(row=ptr, column=1, value="합계").font = Font(bold=True)
c = ws3.cell(row=ptr, column=2, value=sum(d[1] for d in datasets))
c.font = Font(bold=True)
c.number_format = num_fmt
c = ws3.cell(row=ptr, column=3, value=sum(d[2] for d in datasets))
c.font = Font(bold=True)
c.number_format = num_fmt
c = ws3.cell(row=ptr, column=4, value=sum(d[3] for d in datasets))
c.number_format = "$#,##0.00"
c.font = Font(bold=True, size=12, color="C00000")

ws3["A34"] = "※ 현재 잔액 $5.40으로는 전체 재인제스트 불가 — 추가 크레딧 $50+ 필요"
ws3["A34"].font = Font(size=10, bold=True, color="C00000")
ws3["A35"] = "※ 내일 미팅(3/3)에서 POC용 크레딧 추가 제공 요청 필요"
ws3["A35"].font = Font(size=10, color="2F5496")
ws3["A36"] = "※ HWP 파일 직접 지원 여부 미확인 — PDF 변환 필요 시 추가 공수 발생"
ws3["A36"].font = Font(size=9, color="666666")

ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 12
ws3.column_dimensions["E"].width = 45

# Save
output = Path("data/업스테이지_OCR_테스트_리포트_20260302.xlsx")
wb.save(output)
print(f"저장 완료: {output.resolve()}")
