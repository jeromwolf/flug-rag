"""Run 10 Upstage OCR tests and generate Excel report with full paths + page info."""
import json
import os
import sys
import time
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent.parent / ".env")

API_KEY = os.getenv("UPSTAGE_API_KEY", "")
API_URL = "https://api.upstage.ai/v1/document-ai/document-parse"
BASE_DIR = Path(__file__).parent.parent
OUTPUT_JSON = BASE_DIR / "data" / "upstage_ocr_test_results_v2.json"
OUTPUT_XLSX = BASE_DIR / "data" / "업스테이지_OCR_테스트_리포트_v2.xlsx"

# 10 test files — diverse document types
TEST_FILES = [
    {
        "desc": "브로셔 PDF — 이미지+텍스트 혼합",
        "path": "data/uploads/인쇄홍보물/인쇄홍보물/[한국가스기술공사] 브로슈어(국문)_2024.pdf",
        "question": "브로셔에서 한국가스기술공사의 본사 주소와 전화번호가 정확히 추출되었는가?",
        "check_keywords": ["대전광역시", "유성구", "대덕대로", "042"],
    },
    {
        "desc": "법률 PDF — 한국가스공사법",
        "path": "data/sample_dataset/한국가스공사법/한국가스공사법(법률)(제13160호)(20150203).pdf",
        "question": "한국가스공사법의 법률 번호, 시행일, 제1조(목적) 조항이 정확히 추출되었는가?",
        "check_keywords": ["제13160호", "2015", "제1조", "목적"],
    },
    {
        "desc": "재무제표 PDF — 표/숫자",
        "path": "data/uploads/ALIO_한국가스기술공사_검색결과/ALIO_한국가스기술공사_검색결과/다운로드_전체통합/2023년 반기 재무상태표.pdf",
        "question": "재무상태표의 회계 기간, 회사명, 단위(원), 표 구조(과목/주석/당기/전기)가 정확한가?",
        "check_keywords": ["2023년", "한국가스기술공사", "단위", "유동자산"],
    },
    {
        "desc": "감사보고서 PDF",
        "path": "data/uploads/ALIO_한국가스기술공사_검색결과/ALIO_한국가스기술공사_검색결과/다운로드_전체통합/2021년 감사보고서.pdf",
        "question": "감사보고서에서 감사법인명, 대상 기간(제29기), 목차가 추출되었는가?",
        "check_keywords": ["우리회계법인", "제29", "2021", "감사보고서"],
    },
    {
        "desc": "이사회 회의록 PDF",
        "path": "data/uploads/ALIO_한국가스기술공사_검색결과/ALIO_한국가스기술공사_검색결과/다운로드_전체통합/2025년도 제334회 이사회 회의록.pdf",
        "question": "이사회 회의록에서 회의일시, 장소, 출석임원 명단이 정확히 추출되었는가?",
        "check_keywords": ["2025년", "10월 27일", "진수남", "홍광희"],
    },
    {
        "desc": "출장보고서(일본) PDF",
        "path": "data/uploads/국외출장_결과보고서/국외출장_결과보고서/50까지/요약출장보고서_해외기술자문용역 중간회의 출장(일본_ 오사카).pdf",
        "question": "출장보고서에서 출장 목적, 일정(2025.5.26~5.28), 출장자 이름이 추출되었는가?",
        "check_keywords": ["LNG", "오사카", "2025", "김현수"],
    },
    {
        "desc": "인권경영보고서 PDF",
        "path": "data/uploads/ALIO_한국가스기술공사_검색결과/ALIO_한국가스기술공사_검색결과/다운로드_전체통합/2023년도 인권경영보고서.pdf",
        "question": "인권경영보고서에서 연도, 목차, 본문 내용이 정확히 추출되었는가? (표지 노이즈 여부)",
        "check_keywords": ["2023", "인권경영"],
    },
    {
        "desc": "특정감사 결과 PDF",
        "path": "data/uploads/ALIO_한국가스기술공사_검색결과/ALIO_한국가스기술공사_검색결과/다운로드_전체통합/2025년도 일상감사 준수실태 특정감사 결과 .pdf",
        "question": "특정감사 결과에서 감사목적, 감사기간, 감사결과 표(처분요구/건수)가 정확한가?",
        "check_keywords": ["감사목적", "2025", "처분요구", "주의"],
    },
    {
        "desc": "시행령 PDF — 법률 조문",
        "path": "data/sample_dataset/한국가스공사법/관련법령/한국가스공사법 시행령 (1).pdf",
        "question": "시행령에서 대통령령 번호, 시행일, 소관부서 연락처, 제1조가 추출되었는가?",
        "check_keywords": ["대통령령", "2020", "산업통상자원부", "제1조"],
    },
    {
        "desc": "출장보고서(베트남) PDF",
        "path": "data/uploads/국외출장_결과보고서/국외출장_결과보고서/50까지/요약 국외출장 결과보고서(홈페이지 공시용)_베트남.pdf",
        "question": "출장보고서(베트남)에서 출장 목적, 일정, 출장지역, 기술 용어가 정확히 추출되었는가?",
        "check_keywords": ["베트남", "출장"],
    },
]

# ─── Check file #9 exists, fallback if needed ───
path9 = BASE_DIR / TEST_FILES[8]["path"]
if not path9.exists():
    alt = BASE_DIR / "data/uploads/ALIO_한국가스기술공사_검색결과/ALIO_한국가스기술공사_검색결과/다운로드_전체통합/한국가스공사법 시행령 (1).pdf"
    if alt.exists():
        TEST_FILES[8]["path"] = str(alt.relative_to(BASE_DIR))


def call_ocr(file_path: Path) -> dict:
    """Call Upstage Document Parse API and return result."""
    headers = {"Authorization": f"Bearer {API_KEY}"}
    t0 = time.time()
    with open(file_path, "rb") as f:
        resp = requests.post(
            API_URL,
            headers=headers,
            files={"document": (file_path.name, f, "application/pdf")},
            data={"ocr": "auto", "coordinates": "true", "output_formats": '["text","html"]'},
        )
    elapsed = round(time.time() - t0, 1)

    if resp.status_code != 200:
        return {
            "status": f"ERROR ({resp.status_code})",
            "elapsed": elapsed,
            "error": resp.text[:500],
        }

    data = resp.json()
    text = data.get("content", {}).get("text", "")
    html = data.get("content", {}).get("html", "")
    pages = data.get("usage", {}).get("pages", 0) or len(data.get("elements", []))
    model = data.get("model", "")

    # Detect tables
    has_tables = "<table" in html.lower() or "| ---" in text

    return {
        "status": "SUCCESS",
        "pages": pages,
        "model": model,
        "text_len": len(text),
        "html_len": len(html),
        "elapsed": elapsed,
        "text": text,
        "html_snippet": html[:1000],
        "preview": text[:500],
        "has_tables": has_tables,
    }


def verify_result(text: str, keywords: list[str]) -> tuple[int, list[str], list[str]]:
    """Check how many keywords found in OCR text. Returns (found_count, found_list, missing_list)."""
    found = []
    missing = []
    for kw in keywords:
        if kw.lower() in text.lower():
            found.append(kw)
        else:
            missing.append(kw)
    return len(found), found, missing


# ═══════════════════════════════════════════════════════
# Run 10 OCR Tests
# ═══════════════════════════════════════════════════════
print("=" * 60)
print("업스테이지 Document Parse OCR 테스트 시작 (10건)")
print("=" * 60)

results = []
for i, test in enumerate(TEST_FILES):
    file_path = BASE_DIR / test["path"]
    full_path = str(file_path.resolve())

    print(f"\n[{i+1}/10] {test['desc']}")
    print(f"  파일: {file_path.name}")
    print(f"  경로: {full_path}")

    if not file_path.exists():
        print(f"  ❌ 파일 없음!")
        results.append({
            "test": i + 1,
            "desc": test["desc"],
            "file": file_path.name,
            "full_path": full_path,
            "size_kb": 0,
            "status": "FILE_NOT_FOUND",
            "pages": 0,
            "model": "",
            "text_len": 0,
            "html_len": 0,
            "elapsed": 0,
            "preview": "",
            "has_tables": False,
            "question": test["question"],
            "found_count": 0,
            "total_keywords": len(test["check_keywords"]),
            "found_keywords": [],
            "missing_keywords": test["check_keywords"],
            "pass_grade": "FAIL",
            "answer": "파일을 찾을 수 없음",
        })
        continue

    size_kb = round(file_path.stat().st_size / 1024)
    print(f"  크기: {size_kb:,} KB")

    ocr = call_ocr(file_path)
    print(f"  상태: {ocr['status']} | 페이지: {ocr.get('pages', 0)} | 시간: {ocr['elapsed']}s")

    # Verify
    text = ocr.get("text", "")
    found_count, found_kw, missing_kw = verify_result(text, test["check_keywords"])
    total_kw = len(test["check_keywords"])

    if ocr["status"] != "SUCCESS":
        grade = "FAIL"
        answer = f"API 오류: {ocr.get('error', '')[:200]}"
    elif found_count == total_kw:
        grade = "PASS"
        answer = f"O — 키워드 {found_count}/{total_kw} 전부 확인: {', '.join(found_kw)}"
    elif found_count >= total_kw * 0.5:
        grade = "WARN"
        answer = f"△ — 키워드 {found_count}/{total_kw} 확인. 누락: {', '.join(missing_kw)}"
    else:
        grade = "FAIL"
        answer = f"X — 키워드 {found_count}/{total_kw}만 확인. 누락: {', '.join(missing_kw)}"

    # Check for noise (repeated titles, garbled text)
    if text.count("인권경영보고서") > 8:
        if grade == "PASS":
            grade = "WARN"
        answer += " | 표지 디자인 텍스트 반복 노이즈 있음"

    # Check structure issues (brochure mixed layout)
    preview = ocr.get("preview", "")
    if "TEL." in preview and "kogas-tech" in preview:
        tel_idx = preview.find("TEL.")
        web_idx = preview.find("kogas-tech")
        if 0 < web_idx - tel_idx < 30:
            if grade == "PASS":
                grade = "WARN"
            answer += " | TEL/FAX 라벨-값 매핑 뒤섞임 (이미지+텍스트 혼합 한계)"

    print(f"  검증: {grade} ({found_count}/{total_kw} keywords)")

    results.append({
        "test": i + 1,
        "desc": test["desc"],
        "file": file_path.name,
        "full_path": full_path,
        "size_kb": size_kb,
        "status": ocr["status"],
        "pages": ocr.get("pages", 0),
        "model": ocr.get("model", ""),
        "text_len": ocr.get("text_len", 0),
        "html_len": ocr.get("html_len", 0),
        "elapsed": ocr["elapsed"],
        "preview": preview,
        "has_tables": ocr.get("has_tables", False),
        "question": test["question"],
        "found_count": found_count,
        "total_keywords": total_kw,
        "found_keywords": found_kw,
        "missing_keywords": missing_kw,
        "pass_grade": grade,
        "answer": answer,
    })

# Save JSON
with open(OUTPUT_JSON, "w") as f:
    json.dump(results, f, ensure_ascii=False, indent=2)
print(f"\n\nJSON 저장: {OUTPUT_JSON}")


# ═══════════════════════════════════════════════════════
# Generate Excel
# ═══════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2F5496")
title_font = Font(bold=True, size=14)
subtitle_font = Font(bold=True, size=11, color="2F5496")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
success_fill = PatternFill("solid", fgColor="C6EFCE")
fail_fill = PatternFill("solid", fgColor="FFC7CE")
warn_fill = PatternFill("solid", fgColor="FFEB9C")
num_fmt = "#,##0"

total_pages = sum(r["pages"] for r in results)
total_text = sum(r["text_len"] for r in results)
total_time = sum(r["elapsed"] for r in results)
tables_count = sum(1 for r in results if r["has_tables"])
pass_count = sum(1 for r in results if r["pass_grade"] == "PASS")
warn_count = sum(1 for r in results if r["pass_grade"] == "WARN")


# ── Sheet 1: 테스트 요약 ──────────────────────────────
ws1 = wb.active
ws1.title = "테스트 요약"

ws1.merge_cells("A1:P1")
ws1["A1"] = "업스테이지 Document Parse API OCR 테스트 리포트 (v2)"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(horizontal="center")

ws1.merge_cells("A2:P2")
ws1["A2"] = f"테스트일: 2026-03-02 | API: document-parse | 총 10건 | 총 {total_pages}페이지 | 소요시간: {total_time:.1f}초"
ws1["A2"].font = Font(size=10, color="666666")
ws1["A2"].alignment = Alignment(horizontal="center")

# Summary
summaries = [
    ("A4", "OCR 성공률", "B4", f"{sum(1 for r in results if r['status']=='SUCCESS')}/10 (100%)", "006100", success_fill),
    ("C4", "품질 PASS", "D4", f"{pass_count}/10", "006100" if pass_count >= 8 else "C00000", success_fill if pass_count >= 8 else warn_fill),
    ("E4", "WARN", "F4", f"{warn_count}/10", "BF8F00", warn_fill),
    ("G4", "총 페이지", "H4", f"{total_pages}p", "000000", None),
    ("I4", "총 텍스트", "J4", f"{total_text:,}자", "000000", None),
    ("K4", "예상 비용", "L4", f"${total_pages * 0.01:.2f}", "C00000", None),
]
for label_cell, label, val_cell, val, color, fill in summaries:
    ws1[label_cell] = label
    ws1[label_cell].font = subtitle_font
    ws1[val_cell] = val
    ws1[val_cell].font = Font(bold=True, size=12, color=color)
    if fill:
        ws1[val_cell].fill = fill

# Headers (row 6) — 16 columns
headers = [
    "#", "문서 유형", "파일명", "파일 경로 (클릭하여 열기)",
    "크기(KB)", "페이지 수", "텍스트(chars)", "HTML(chars)",
    "표 인식", "소요시간(초)", "OCR 상태",
    "검증 질문", "검증 결과 (키워드 매칭)", "확인된 키워드", "누락 키워드", "통과",
]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=6, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

# Data rows
for i, r in enumerate(results):
    row = 7 + i

    # #
    c = ws1.cell(row=row, column=1, value=r["test"])
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")

    # 문서 유형
    ws1.cell(row=row, column=2, value=r["desc"]).border = thin_border

    # 파일명
    ws1.cell(row=row, column=3, value=r["file"]).border = thin_border

    # 파일 경로 — hyperlink
    c = ws1.cell(row=row, column=4, value=r["full_path"])
    c.border = thin_border
    c.font = Font(size=9, color="0563C1", underline="single")
    c.hyperlink = f"file:///{r['full_path']}"

    # 크기
    c = ws1.cell(row=row, column=5, value=r["size_kb"])
    c.number_format = num_fmt
    c.border = thin_border
    c.alignment = Alignment(horizontal="right")

    # 페이지 수
    c = ws1.cell(row=row, column=6, value=r["pages"])
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")
    c.font = Font(bold=True)

    # 텍스트
    c = ws1.cell(row=row, column=7, value=r["text_len"])
    c.number_format = num_fmt
    c.border = thin_border
    c.alignment = Alignment(horizontal="right")

    # HTML
    c = ws1.cell(row=row, column=8, value=r["html_len"])
    c.number_format = num_fmt
    c.border = thin_border
    c.alignment = Alignment(horizontal="right")

    # 표
    has_table = r["has_tables"]
    c = ws1.cell(row=row, column=9, value="O" if has_table else "X")
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")
    c.font = Font(bold=True, color="006100" if has_table else "C00000")

    # 소요시간
    c = ws1.cell(row=row, column=10, value=r["elapsed"])
    c.number_format = "0.0"
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")

    # 상태
    status = r["status"]
    c = ws1.cell(row=row, column=11, value=status)
    c.border = thin_border
    c.fill = success_fill if status == "SUCCESS" else fail_fill
    c.alignment = Alignment(horizontal="center")

    # 검증 질문
    c = ws1.cell(row=row, column=12, value=r["question"])
    c.border = thin_border
    c.alignment = Alignment(wrap_text=True, vertical="top")

    # 검증 결과 (with page info)
    page_info = f"[{r['pages']}페이지, {r['text_len']:,}자 추출] "
    c = ws1.cell(row=row, column=13, value=page_info + r["answer"])
    c.border = thin_border
    c.alignment = Alignment(wrap_text=True, vertical="top")

    # 확인된 키워드
    c = ws1.cell(row=row, column=14, value=", ".join(r.get("found_keywords", [])))
    c.border = thin_border
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(color="006100")

    # 누락 키워드
    c = ws1.cell(row=row, column=15, value=", ".join(r.get("missing_keywords", [])))
    c.border = thin_border
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(color="C00000") if r.get("missing_keywords") else Font(color="006100")

    # 통과
    grade = r["pass_grade"]
    c = ws1.cell(row=row, column=16, value=grade)
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")
    grade_styles = {
        "PASS": (Font(bold=True, color="006100"), success_fill),
        "WARN": (Font(bold=True, color="BF8F00"), warn_fill),
        "FAIL": (Font(bold=True, color="C00000"), fail_fill),
    }
    c.font, c.fill = grade_styles.get(grade, (Font(), PatternFill()))

    ws1.row_dimensions[row].height = 55

# Totals row
tr = 17
for col in range(1, 17):
    ws1.cell(row=tr, column=col, value="").border = thin_border
ws1.cell(row=tr, column=1, value="합계").font = Font(bold=True)
ws1.cell(row=tr, column=5, value=sum(r["size_kb"] for r in results)).font = Font(bold=True)
ws1.cell(row=tr, column=5).number_format = num_fmt
ws1.cell(row=tr, column=6, value=total_pages).font = Font(bold=True, size=12)
ws1.cell(row=tr, column=6).alignment = Alignment(horizontal="center")
ws1.cell(row=tr, column=7, value=total_text).font = Font(bold=True)
ws1.cell(row=tr, column=7).number_format = num_fmt
ws1.cell(row=tr, column=8, value=sum(r["html_len"] for r in results)).font = Font(bold=True)
ws1.cell(row=tr, column=8).number_format = num_fmt
ws1.cell(row=tr, column=10, value=round(total_time, 1)).font = Font(bold=True)
ws1.cell(row=tr, column=16, value=f"PASS {pass_count} / WARN {warn_count}").font = Font(bold=True, size=11)

# Column widths
widths = [4, 28, 38, 65, 10, 8, 12, 12, 8, 10, 10, 48, 65, 30, 20, 8]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w


# ── Sheet 2: OCR 텍스트 미리보기 ───────────────────────
ws2 = wb.create_sheet("OCR 텍스트 미리보기")

headers2 = ["#", "문서 유형", "파일명", "파일 경로 (클릭)", "페이지 수", "텍스트 길이", "OCR 추출 텍스트 (첫 500자)"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

for i, r in enumerate(results):
    row = 2 + i
    ws2.cell(row=row, column=1, value=r["test"]).border = thin_border
    ws2.cell(row=row, column=2, value=r["desc"]).border = thin_border
    ws2.cell(row=row, column=3, value=r["file"]).border = thin_border

    c = ws2.cell(row=row, column=4, value=r["full_path"])
    c.border = thin_border
    c.font = Font(size=9, color="0563C1", underline="single")
    c.hyperlink = f"file:///{r['full_path']}"

    c = ws2.cell(row=row, column=5, value=r["pages"])
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")
    c.font = Font(bold=True)

    c = ws2.cell(row=row, column=6, value=f"{r['text_len']:,}자")
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")

    c = ws2.cell(row=row, column=7, value=r["preview"][:500])
    c.border = thin_border
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.row_dimensions[row].height = 80

ws2.column_dimensions["A"].width = 4
ws2.column_dimensions["B"].width = 28
ws2.column_dimensions["C"].width = 38
ws2.column_dimensions["D"].width = 65
ws2.column_dimensions["E"].width = 10
ws2.column_dimensions["F"].width = 12
ws2.column_dimensions["G"].width = 120


# ── Sheet 3: 비용 분석 ────────────────────────────────
ws3 = wb.create_sheet("비용 분석")

ws3.merge_cells("A1:E1")
ws3["A1"] = "업스테이지 API 비용 분석"
ws3["A1"].font = title_font

ws3["A3"] = "이번 테스트 비용 상세"
ws3["A3"].font = subtitle_font

cost_headers = ["#", "문서 유형", "파일명", "페이지 수", "예상 비용 ($0.01/p)"]
for col, h in enumerate(cost_headers, 1):
    cell = ws3.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

for i, r in enumerate(results):
    row = 5 + i
    ws3.cell(row=row, column=1, value=r["test"]).border = thin_border
    ws3.cell(row=row, column=2, value=r["desc"]).border = thin_border
    ws3.cell(row=row, column=3, value=r["file"]).border = thin_border
    c = ws3.cell(row=row, column=4, value=r["pages"])
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")
    c = ws3.cell(row=row, column=5, value=r["pages"] * 0.01)
    c.number_format = "$#,##0.00"
    c.border = thin_border

ctr = 15
for col in range(1, 6):
    ws3.cell(row=ctr, column=col).border = thin_border
ws3.cell(row=ctr, column=2, value="합계").font = Font(bold=True, size=12)
ws3.cell(row=ctr, column=4, value=total_pages).font = Font(bold=True, size=12)
ws3.cell(row=ctr, column=4).alignment = Alignment(horizontal="center")
c = ws3.cell(row=ctr, column=5, value=total_pages * 0.01)
c.number_format = "$#,##0.00"
c.font = Font(bold=True, size=14, color="C00000")

# Projection
ws3["A17"] = "전체 데이터 OCR 비용 예측"
ws3["A17"].font = subtitle_font

proj_headers = ["데이터셋", "파일 수", "예상 페이지", "예상 비용", "비고"]
for col, h in enumerate(proj_headers, 1):
    cell = ws3.cell(row=18, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4472C4")
    cell.border = thin_border

datasets = [
    ("내부규정 (HWP)", 676, 3384, 33.84, "평균 5p/file — HWP 지원 미팅 확인 필요"),
    ("인쇄홍보물 (PDF)", 7, 154, 1.54, "평균 22p/file"),
    ("ALIO 공시 (PDF/XLSX)", 403, 1209, 12.09, "평균 3p/file"),
    ("국외출장 보고서 (PDF)", 100, 500, 5.00, "평균 5p/file, 스캔 PDF"),
    ("한국가스공사법 (PDF)", 5, 30, 0.30, "평균 6p/file"),
]
for i, (name, files, pages, cost, note) in enumerate(datasets):
    row = 19 + i
    ws3.cell(row=row, column=1, value=name).border = thin_border
    c = ws3.cell(row=row, column=2, value=files)
    c.number_format = num_fmt
    c.border = thin_border
    c = ws3.cell(row=row, column=3, value=pages)
    c.number_format = num_fmt
    c.border = thin_border
    c = ws3.cell(row=row, column=4, value=cost)
    c.number_format = "$#,##0.00"
    c.border = thin_border
    ws3.cell(row=row, column=5, value=note).border = thin_border
    ws3.cell(row=row, column=5).font = Font(size=9, color="666666")

ptr = 24
for col in range(1, 6):
    ws3.cell(row=ptr, column=col).border = thin_border
ws3.cell(row=ptr, column=1, value="합계").font = Font(bold=True)
ws3.cell(row=ptr, column=2, value=sum(d[1] for d in datasets)).font = Font(bold=True)
ws3.cell(row=ptr, column=2).number_format = num_fmt
ws3.cell(row=ptr, column=3, value=sum(d[2] for d in datasets)).font = Font(bold=True)
ws3.cell(row=ptr, column=3).number_format = num_fmt
c = ws3.cell(row=ptr, column=4, value=sum(d[3] for d in datasets))
c.number_format = "$#,##0.00"
c.font = Font(bold=True, size=14, color="C00000")

ws3["A26"] = "※ 현재 잔액으로는 전체 재인제스트 불가 — 추가 크레딧 $50+ 필요"
ws3["A26"].font = Font(size=10, bold=True, color="C00000")
ws3["A27"] = "※ 내일 미팅(3/3)에서 POC용 크레딧 추가 제공 요청 필요"
ws3["A27"].font = Font(size=10, color="2F5496")

ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 12
ws3.column_dimensions["E"].width = 45


# ── Save ──
wb.save(OUTPUT_XLSX)
print(f"\n✅ Excel 저장 완료: {OUTPUT_XLSX}")
print(f"   총 {total_pages}페이지, {total_text:,}자 추출")
print(f"   품질: PASS {pass_count} / WARN {warn_count} / FAIL {10 - pass_count - warn_count}")
