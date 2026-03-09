"""Test golden dataset samples against RunPod RAG API and export to Excel.

Evaluation logic:
  - Extract key terms (2+ char Korean words) from expected_answer
  - Check overlap ratio with actual_answer
  - Detect false negatives: answer says "없습니다" when expected has real content
  - Detect false positives: answer has specific content when expected says "없습니다"
  - OK: keyword overlap >= 30% OR both agree on negative
  - MISS: keyword overlap < 30% OR false negative/positive detected
"""
import json
import re
import sys
import time
import urllib.request
import urllib.error
from pathlib import Path

API_BASE = "https://7rzubyo9fsfmco-8000.proxy.runpod.net"
DATASET_PATH = Path(__file__).parent.parent / "tests" / "golden_dataset_evaluation.json"
OUTPUT_EXCEL = Path(__file__).parent.parent / "data" / "golden_sample_test_results.xlsx"

# ---------------------------------------------------------------------------
# Evaluation helpers
# ---------------------------------------------------------------------------

# Phrases that indicate "no relevant regulation found"
NEGATIVE_PHRASES = ["없습니다", "확인되지", "찾을 수 없", "관련 규정이 없", "해당 규정은 없",
                    "다루는 규정을 찾을 수 없", "관련된 규정이 없", "포함되어 있지 않"]


def extract_keywords(text: str) -> set[str]:
    """Extract meaningful Korean keywords (2+ chars) from text."""
    # Remove common filler/stop words
    stop = {"있습니다", "없습니다", "됩니다", "합니다", "입니다", "것입니다", "대한",
            "위한", "따른", "의한", "에서", "으로", "에는", "하는", "되는", "있는",
            "따라", "통해", "관한", "대해", "것이", "수도", "또는", "그리고", "이를"}
    words = set(re.findall(r'[가-힣]{2,}', text))
    return words - stop


def is_negative_answer(text: str) -> bool:
    """Check if the answer essentially says 'no relevant info found'."""
    return any(phrase in text for phrase in NEGATIVE_PHRASES)


def evaluate_answer(expected: str, actual: str, category: str) -> tuple[str, float, str]:
    """Evaluate actual answer against expected answer.

    Returns (status, overlap_ratio, reason).
    """
    if not actual or len(actual) < 5:
        return "MISS", 0.0, "답변 없음/너무 짧음"

    expected_negative = is_negative_answer(expected)
    actual_negative = is_negative_answer(actual)

    # Case 1: Both agree it's negative → OK for negative category
    if expected_negative and actual_negative:
        return "OK", 1.0, "부정 답변 일치"

    # Case 2: False negative — expected has content but answer says "없습니다"
    if not expected_negative and actual_negative:
        return "MISS", 0.0, "오답: 실제 답변이 '없습니다'인데 기대 답변에 내용 있음"

    # Case 3: False positive — expected says "없습니다" but answer has content
    if expected_negative and not actual_negative:
        # This could be OK if the answer provides relevant info, but for golden dataset it's wrong
        return "MISS", 0.0, "오답: 기대 답변은 '없습니다'인데 실제 답변에 내용 있음"

    # Case 4: Both have content — check keyword overlap
    expected_kw = extract_keywords(expected)
    actual_kw = extract_keywords(actual)

    if not expected_kw:
        # Can't evaluate without keywords
        return "OK" if len(actual) > 20 else "MISS", 0.5, "키워드 추출 불가"

    overlap = expected_kw & actual_kw
    ratio = len(overlap) / len(expected_kw)

    if ratio >= 0.3:
        return "OK", ratio, f"키워드 {len(overlap)}/{len(expected_kw)} 일치 ({ratio:.0%})"
    else:
        return "MISS", ratio, f"키워드 {len(overlap)}/{len(expected_kw)} 일치 ({ratio:.0%}) — 미달"


def api_post(path, data, token=""):
    url = f"{API_BASE}{path}"
    body = json.dumps(data).encode("utf-8")
    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("Content-Type", "application/json")
    req.add_header("User-Agent", "Mozilla/5.0 (golden-test/1.0)")
    if token:
        req.add_header("Authorization", f"Bearer {token}")
    try:
        with urllib.request.urlopen(req, timeout=180) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body_text = e.read().decode("utf-8", errors="replace")
        print(f"  HTTP {e.code}: {body_text[:200]}", file=sys.stderr)
        return None
    except Exception as e:
        print(f"  Error: {e}", file=sys.stderr)
        return None


def main():
    print("=== Golden Dataset Sample Test (10 questions) ===\n", flush=True)

    # Login
    print("Logging in...", flush=True)
    resp = api_post("/api/auth/login", {"username": "admin", "password": "admin123"})
    if not resp or "access_token" not in resp:
        print("Login failed!", flush=True)
        sys.exit(1)
    token = resp["access_token"]
    print("Login OK.\n", flush=True)

    # Load dataset
    with open(DATASET_PATH, "r", encoding="utf-8") as f:
        dataset = json.load(f)
    questions = dataset["questions"]

    # Select 10 samples: 4 factual, 2 multi_hop, 2 inference, 2 negative
    samples = []
    targets = {"factual": 4, "multi_hop": 2, "inference": 2, "negative": 2}
    for cat, n in targets.items():
        pool = [q for q in questions if q["category"] == cat]
        samples.extend(pool[:n])

    print(f"Selected {len(samples)} questions\n", flush=True)

    header = f"{'#':>3} {'Category':<10} {'Status':<6} {'Ovlp':>5} {'Time':>6} {'Conf':>5} {'Src':>3} Question"
    print(header, flush=True)
    print("-" * 110, flush=True)

    results = []
    for i, q in enumerate(samples):
        start = time.time()
        resp = api_post("/api/chat", {
            "message": q["question"],
            "temperature": 0.1,
        }, token=token)
        elapsed = time.time() - start

        if resp:
            answer = resp.get("content", resp.get("answer", ""))
            sources = resp.get("sources", [])
            confidence = resp.get("confidence", 0)
            model = resp.get("model", "")

            # Evaluate answer quality
            status, overlap, reason = evaluate_answer(
                q["expected_answer"], answer, q["category"],
            )

            results.append({
                "id": q["id"],
                "category": q["category"],
                "difficulty": q["difficulty"],
                "question": q["question"],
                "expected_answer": q["expected_answer"],
                "actual_answer": answer,
                "status": status,
                "overlap": round(overlap, 2),
                "eval_reason": reason,
                "confidence": confidence,
                "sources_count": len(sources),
                "sources": "; ".join([s.get("source", s.get("filename", "")) for s in sources][:3]) if sources else "",
                "response_time": round(elapsed, 1),
                "model": model,
            })
        else:
            status = "ERR"
            results.append({
                "id": q["id"],
                "category": q["category"],
                "difficulty": q["difficulty"],
                "question": q["question"],
                "expected_answer": q["expected_answer"],
                "actual_answer": "(Error)",
                "status": "ERR",
                "overlap": 0.0,
                "eval_reason": "API 에러",
                "confidence": 0,
                "sources_count": 0,
                "sources": "",
                "response_time": round(elapsed, 1),
                "model": "",
            })

        r = results[-1]
        short_q = q["question"][:40]
        print(f"{i+1:3d} {r['category']:<10} {r['status']:<6} {r['overlap']:5.0%} {r['response_time']:5.1f}s {r['confidence']:5.2f} {r['sources_count']:3d} {short_q}", flush=True)
        if r["status"] != "OK":
            print(f"    → {r['eval_reason']}", flush=True)
        time.sleep(0.3)

    # Summary
    ok = sum(1 for r in results if r["status"] == "OK")
    total = len(results)
    avg_time = sum(r["response_time"] for r in results) / max(total, 1)
    avg_conf = sum(r["confidence"] for r in results) / max(total, 1)

    print(f"\n{'='*60}", flush=True)
    print(f"Results: {ok}/{total} OK ({ok/total*100:.0f}%)", flush=True)
    print(f"Avg time: {avg_time:.1f}s | Avg confidence: {avg_conf:.2f}", flush=True)
    print(f"{'='*60}\n", flush=True)

    # Export to Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "골든 데이터셋 테스트 결과"

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        miss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        err_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Headers
        headers = ["#", "카테고리", "난이도", "질문", "기대 답변", "실제 답변", "결과",
                    "일치율", "평가사유", "신뢰도", "소스수", "소스 문서", "응답시간(s)", "모델"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        # Data rows
        for row_idx, r in enumerate(results, 2):
            values = [r["id"], r["category"], r["difficulty"], r["question"],
                      r["expected_answer"], r["actual_answer"], r["status"],
                      f"{r['overlap']:.0%}", r["eval_reason"],
                      r["confidence"], r["sources_count"], r["sources"],
                      r["response_time"], r["model"]]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Color status cell
            status_cell = ws.cell(row=row_idx, column=7)
            if r["status"] == "OK":
                status_cell.fill = ok_fill
            elif r["status"] == "MISS":
                status_cell.fill = miss_fill
            else:
                status_cell.fill = err_fill

        # Summary row
        sum_row = len(results) + 3
        ws.cell(row=sum_row, column=1, value="요약").font = Font(bold=True, size=12)
        ws.cell(row=sum_row + 1, column=1, value=f"총 문항: {total}")
        ws.cell(row=sum_row + 2, column=1, value=f"성공: {ok}/{total} ({ok/total*100:.0f}%)")
        ws.cell(row=sum_row + 3, column=1, value=f"평균 응답시간: {avg_time:.1f}s")
        ws.cell(row=sum_row + 4, column=1, value=f"평균 신뢰도: {avg_conf:.2f}")

        # Column widths
        widths = [5, 12, 8, 45, 45, 60, 8, 8, 6, 40, 10, 25]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        wb.save(str(OUTPUT_EXCEL))
        print(f"Excel saved: {OUTPUT_EXCEL}", flush=True)

    except ImportError:
        print("openpyxl not installed, saving as JSON instead", flush=True)
        json_path = OUTPUT_EXCEL.with_suffix(".json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump({"summary": {"total": total, "ok": ok, "avg_time": avg_time},
                        "results": results}, f, ensure_ascii=False, indent=2)
        print(f"JSON saved: {json_path}", flush=True)


if __name__ == "__main__":
    main()
