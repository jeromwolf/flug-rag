"""Excel document loader using openpyxl."""

import asyncio
from pathlib import Path

from openpyxl import load_workbook

from .base import BaseLoader, LoadedDocument


class ExcelLoader(BaseLoader):
    """Load Excel files using openpyxl."""

    supported_extensions = [".xlsx", ".xls"]

    async def load(self, file_path: str | Path) -> LoadedDocument:
        path = self._validate_file(file_path)
        return await asyncio.to_thread(self._load_sync, path)

    def _load_sync(self, path: Path) -> LoadedDocument:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        sheet_names = list(wb.sheetnames)
        all_text_parts = []
        pages = []

        for idx, sheet_name in enumerate(sheet_names):
            ws = wb[sheet_name]
            sheet_text = self._extract_sheet(ws)
            if sheet_text.strip():
                pages.append({
                    "page_num": idx + 1,
                    "content": sheet_text.strip(),
                    "sheet_name": sheet_name,
                })
                all_text_parts.append(f"[시트: {sheet_name}]\n{sheet_text.strip()}")

        wb.close()

        return LoadedDocument(
            content="\n\n".join(all_text_parts),
            metadata={
                "filename": path.name,
                "file_type": "xlsx",
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
            },
            pages=pages,
        )

    def _extract_sheet(self, ws) -> str:
        """Extract sheet data as text table."""
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))
        return "\n".join(rows)
