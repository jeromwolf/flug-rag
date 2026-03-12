"""File parser tool for CSV, Excel, and JSON file analysis."""

import csv
import json
from pathlib import Path
from typing import Any

from agent.mcp.tools.base import (
    BaseTool, ToolDefinition, ToolParameter, ToolParamType, ToolResult,
)


class FileParserTool(BaseTool):
    """파일 파싱 도구 -- CSV, Excel, JSON 파일 분석"""

    def get_definition(self) -> ToolDefinition:
        return ToolDefinition(
            name="file_parser",
            description="CSV, Excel, JSON 파일을 분석합니다.",
            category="data",
            help_text=(
                "CSV, Excel(.xlsx), JSON 파일을 읽고 분석합니다.\n"
                "작업 유형:\n"
                "  - summary: 행 수, 컬럼명, 첫 3행 미리보기\n"
                "  - head: 첫 10행 표시\n"
                "  - columns: 컬럼명 및 데이터 타입\n"
                "  - stats: 컬럼별 기본 통계 (고유값 수 등)\n"
                "  - search: 특정 문자열을 포함하는 행 검색\n"
                "파일 경로는 data 디렉토리 기준 상대 경로 또는 절대 경로를 사용합니다."
            ),
            parameters=[
                ToolParameter(
                    name="file_path",
                    type=ToolParamType.STRING,
                    description="파일 경로 (data 디렉토리 기준 상대 경로 또는 절대 경로)",
                ),
                ToolParameter(
                    name="operation",
                    type=ToolParamType.STRING,
                    description="작업 유형",
                    required=False,
                    default="summary",
                    enum=["summary", "head", "columns", "stats", "search"],
                ),
                ToolParameter(
                    name="query",
                    type=ToolParamType.STRING,
                    description="검색 문자열 (search 작업 시 사용)",
                    required=False,
                ),
            ],
        )

    def _resolve_path(self, file_path: str) -> Path:
        """Resolve file path relative to data_dir or as absolute."""
        from config.settings import settings

        path = Path(file_path)
        if path.is_absolute():
            return path
        return Path(settings.data_dir) / file_path

    def _read_csv(self, path: Path) -> list[dict[str, str]]:
        """Read CSV file and return list of row dicts."""
        encodings = ["utf-8", "cp949", "euc-kr"]
        for enc in encodings:
            try:
                with open(path, "r", encoding=enc, newline="") as f:
                    reader = csv.DictReader(f)
                    return list(reader)
            except (UnicodeDecodeError, UnicodeError):
                continue
        raise ValueError(f"Cannot decode CSV file with supported encodings: {encodings}")

    def _read_excel(self, path: Path) -> tuple[list[str], list[list[Any]]]:
        """Read Excel file using openpyxl. Returns (headers, rows)."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is required for Excel file parsing. Install with: pip install openpyxl")

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("Excel file has no active worksheet")

        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            wb.close()
            return [], []

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
        data_rows = []
        for row in rows_iter:
            data_rows.append([cell for cell in row])

        wb.close()
        return headers, data_rows

    def _read_json(self, path: Path) -> Any:
        """Read JSON file."""
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    def _load_data(self, path: Path) -> tuple[list[str], list[dict[str, Any]]]:
        """
        Load data from file. Returns (column_names, list_of_row_dicts).
        """
        suffix = path.suffix.lower()

        if suffix == ".csv":
            rows = self._read_csv(path)
            columns = list(rows[0].keys()) if rows else []
            return columns, rows

        elif suffix in (".xlsx", ".xls"):
            headers, data_rows = self._read_excel(path)
            rows = []
            for row_data in data_rows:
                row_dict = {}
                for i, h in enumerate(headers):
                    val = row_data[i] if i < len(row_data) else None
                    row_dict[h] = val
                rows.append(row_dict)
            return headers, rows

        elif suffix == ".json":
            data = self._read_json(path)
            if isinstance(data, list) and data and isinstance(data[0], dict):
                columns = list(data[0].keys())
                return columns, data
            elif isinstance(data, dict):
                # Single object: wrap as one-row table
                columns = list(data.keys())
                return columns, [data]
            else:
                # Array of primitives or other structure
                return ["value"], [{"value": item} for item in (data if isinstance(data, list) else [data])]

        else:
            raise ValueError(f"Unsupported file type: {suffix}. Supported: .csv, .xlsx, .xls, .json")

    async def execute(self, **kwargs) -> ToolResult:
        file_path = kwargs.get("file_path", "")
        operation = kwargs.get("operation", "summary")
        query = kwargs.get("query", "")

        if not file_path:
            return ToolResult(success=False, error="file_path parameter is required")

        try:
            path = self._resolve_path(file_path)
        except Exception as e:
            return ToolResult(success=False, error=f"Path resolution error: {e}")

        if not path.exists():
            return ToolResult(success=False, error=f"File not found: {path}")
        if not path.is_file():
            return ToolResult(success=False, error=f"Not a file: {path}")

        try:
            columns, rows = self._load_data(path)

            if operation == "summary":
                result = self._summary(columns, rows, path)
            elif operation == "head":
                result = self._head(columns, rows)
            elif operation == "columns":
                result = self._columns(columns, rows)
            elif operation == "stats":
                result = self._stats(columns, rows)
            elif operation == "search":
                if not query:
                    return ToolResult(success=False, error="query parameter is required for search operation")
                result = self._search(columns, rows, query)
            else:
                return ToolResult(success=False, error=f"Unknown operation: {operation}")

            return ToolResult(
                success=True,
                data=result,
                metadata={"file_path": str(path), "operation": operation},
            )

        except ImportError as e:
            return ToolResult(success=False, error=str(e))
        except Exception as e:
            return ToolResult(success=False, error=f"File parsing error: {e}")

    def _summary(self, columns: list[str], rows: list[dict], path: Path) -> dict:
        """Row count, column names, first 3 rows."""
        preview = self._serialize_rows(rows[:3])
        return {
            "file_name": path.name,
            "file_size": path.stat().st_size,
            "row_count": len(rows),
            "column_count": len(columns),
            "columns": columns,
            "preview": preview,
        }

    def _head(self, columns: list[str], rows: list[dict]) -> dict:
        """First 10 rows."""
        return {
            "columns": columns,
            "row_count": len(rows),
            "rows": self._serialize_rows(rows[:10]),
        }

    def _columns(self, columns: list[str], rows: list[dict]) -> dict:
        """Column names and inferred types."""
        col_info = []
        for col in columns:
            sample_values = [r.get(col) for r in rows[:100] if r.get(col) is not None]
            inferred_type = self._infer_type(sample_values)
            col_info.append({
                "name": col,
                "type": inferred_type,
                "non_null_count": len(sample_values),
                "sample": str(sample_values[0]) if sample_values else None,
            })
        return {"columns": col_info, "total_columns": len(columns)}

    def _stats(self, columns: list[str], rows: list[dict]) -> dict:
        """Basic stats per column."""
        stats = {}
        for col in columns:
            values = [r.get(col) for r in rows if r.get(col) is not None]
            unique_values = set(str(v) for v in values)
            stats[col] = {
                "count": len(values),
                "unique": len(unique_values),
                "null_count": len(rows) - len(values),
            }
            # Show top 5 most common values
            from collections import Counter
            str_values = [str(v) for v in values]
            top = Counter(str_values).most_common(5)
            stats[col]["top_values"] = [{"value": v, "count": c} for v, c in top]
        return {"row_count": len(rows), "column_stats": stats}

    def _search(self, columns: list[str], rows: list[dict], query: str) -> dict:
        """Filter rows containing query string in any column."""
        query_lower = query.lower()
        matched = []
        for row in rows:
            for val in row.values():
                if val is not None and query_lower in str(val).lower():
                    matched.append(row)
                    break
            if len(matched) >= 50:  # Limit results
                break

        return {
            "query": query,
            "matched_count": len(matched),
            "total_rows": len(rows),
            "rows": self._serialize_rows(matched[:50]),
        }

    @staticmethod
    def _infer_type(values: list[Any]) -> str:
        """Infer column data type from sample values."""
        if not values:
            return "unknown"

        int_count = 0
        float_count = 0
        for v in values[:50]:
            if isinstance(v, bool):
                return "boolean"
            elif isinstance(v, int):
                int_count += 1
            elif isinstance(v, float):
                float_count += 1
            elif isinstance(v, str):
                try:
                    int(v)
                    int_count += 1
                except ValueError:
                    try:
                        float(v)
                        float_count += 1
                    except ValueError:
                        pass

        total = min(len(values), 50)
        if int_count + float_count > total * 0.7:
            return "float" if float_count > 0 else "integer"
        return "string"

    @staticmethod
    def _serialize_rows(rows: list[dict]) -> list[dict]:
        """Convert row values to JSON-safe types."""
        serialized = []
        for row in rows:
            safe_row = {}
            for k, v in row.items():
                if v is None:
                    safe_row[k] = None
                elif isinstance(v, (int, float, bool, str)):
                    safe_row[k] = v
                else:
                    safe_row[k] = str(v)
            serialized.append(safe_row)
        return serialized
