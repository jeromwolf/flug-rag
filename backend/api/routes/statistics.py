"""
통계 엔진 API
SFR-009: 사용 통계 + 키워드 분석 + Excel 내보내기
"""
import logging
from datetime import datetime, timezone, timedelta
from io import BytesIO
from typing import Annotated, Literal

import aiosqlite
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from auth.dependencies import get_current_user, require_role
from auth.models import Role, User

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/stats", tags=["statistics"])


class UsageStats(BaseModel):
    period: str
    total_sessions: int = 0
    total_queries: int = 0
    unique_users: int = 0
    avg_queries_per_session: float = 0.0
    daily_breakdown: list[dict] = []


class KeywordStats(BaseModel):
    keywords: list[dict]  # [{"keyword": "가스", "count": 42}, ...]
    total_queries: int = 0


class UserUsageStats(BaseModel):
    users: list[dict]  # [{"username": "admin", "query_count": 10, "department": "..."}, ...]
    departments: list[dict]  # [{"department": "안전팀", "query_count": 50}, ...]


async def _get_memory_db_path():
    """Get memory.db path."""
    from config.settings import settings
    return settings.data_dir / "memory.db"


async def _get_audit_db_path():
    """Get audit.db path."""
    from config.settings import settings
    return settings.data_dir / "audit.db"


def _get_period_start(period: str) -> datetime:
    """Calculate period start datetime."""
    now = datetime.now(timezone.utc)
    if period == "day":
        return now - timedelta(days=1)
    elif period == "week":
        return now - timedelta(weeks=1)
    elif period == "month":
        return now - timedelta(days=30)
    return now - timedelta(days=1)


@router.get("/usage", response_model=UsageStats)
async def get_usage_stats(
    current_user: Annotated[User, Depends(require_role([Role.ADMIN, Role.MANAGER]))],
    period: Literal["day", "week", "month"] = "week",
):
    """사용량 통계 (세션수, 질의수, 사용자수)."""
    db_path = await _get_memory_db_path()
    start = _get_period_start(period)
    start_str = start.isoformat()

    async with aiosqlite.connect(db_path) as db:
        db.row_factory = aiosqlite.Row

        # Total sessions in period
        async with db.execute(
            "SELECT COUNT(DISTINCT session_id) as cnt FROM messages WHERE created_at >= ?",
            (start_str,)
        ) as cur:
            row = await cur.fetchone()
            total_sessions = row["cnt"] if row else 0

        # Total user queries (role='user')
        async with db.execute(
            "SELECT COUNT(*) as cnt FROM messages WHERE role='user' AND created_at >= ?",
            (start_str,)
        ) as cur:
            row = await cur.fetchone()
            total_queries = row["cnt"] if row else 0

        # Unique users (from sessions metadata or audit)
        unique_users = 0
        try:
            audit_path = await _get_audit_db_path()
            async with aiosqlite.connect(audit_path) as adb:
                async with adb.execute(
                    "SELECT COUNT(DISTINCT user_id) as cnt FROM audit_log WHERE timestamp >= ?",
                    (start_str,)
                ) as cur2:
                    row2 = await cur2.fetchone()
                    unique_users = row2[0] if row2 else 0
        except Exception:
            pass

        # Daily breakdown
        daily = []
        async with db.execute(
            """SELECT DATE(created_at) as day, COUNT(*) as cnt
               FROM messages WHERE role='user' AND created_at >= ?
               GROUP BY DATE(created_at) ORDER BY day""",
            (start_str,)
        ) as cur:
            async for row in cur:
                daily.append({"date": row["day"], "queries": row["cnt"]})

    avg = round(total_queries / max(total_sessions, 1), 2)

    return UsageStats(
        period=period,
        total_sessions=total_sessions,
        total_queries=total_queries,
        unique_users=unique_users,
        avg_queries_per_session=avg,
        daily_breakdown=daily,
    )


@router.get("/keywords", response_model=KeywordStats)
async def get_keyword_stats(
    current_user: Annotated[User, Depends(require_role([Role.ADMIN, Role.MANAGER]))],
    period: Literal["day", "week", "month"] = "week",
    top_n: Annotated[int, Query(le=100)] = 30,
):
    """키워드/주제어 통계."""
    import re
    from collections import Counter

    db_path = await _get_memory_db_path()
    start = _get_period_start(period)

    async with aiosqlite.connect(db_path) as db:
        async with db.execute(
            "SELECT content FROM messages WHERE role='user' AND created_at >= ?",
            (start.isoformat(),)
        ) as cur:
            rows = await cur.fetchall()

    # Simple Korean keyword extraction (2+ char nouns)
    counter = Counter()
    total = len(rows)
    for row in rows:
        text = row[0] if row[0] else ""
        # Extract Korean words of 2+ chars
        words = re.findall(r'[가-힣]{2,}', text)
        # Filter common stop words
        stop_words = {"입니다", "있는", "하는", "되는", "에서", "으로", "합니다", "것입니다", "무엇", "어떻게", "알려줘", "알려주세요", "해주세요"}
        words = [w for w in words if w not in stop_words]
        counter.update(words)

    keywords = [{"keyword": k, "count": v} for k, v in counter.most_common(top_n)]

    return KeywordStats(keywords=keywords, total_queries=total)


@router.get("/users", response_model=UserUsageStats)
async def get_user_usage_stats(
    current_user: Annotated[User, Depends(require_role([Role.ADMIN]))],
    period: Literal["day", "week", "month"] = "week",
):
    """사용자별/부서별 사용량 통계."""
    from collections import Counter
    audit_path = await _get_audit_db_path()
    start = _get_period_start(period)

    users_data = []
    dept_counter = Counter()

    try:
        async with aiosqlite.connect(audit_path) as db:
            db.row_factory = aiosqlite.Row
            async with db.execute(
                """SELECT user_id, username, action, COUNT(*) as cnt
                   FROM audit_log
                   WHERE timestamp >= ? AND action IN ('LOGIN', 'DOCUMENT_ACCESS')
                   GROUP BY user_id, username, action
                   ORDER BY cnt DESC""",
                (start.isoformat(),)
            ) as cur:
                async for row in cur:
                    users_data.append({
                        "username": row["username"],
                        "user_id": row["user_id"],
                        "action": row["action"],
                        "count": row["cnt"],
                    })
    except Exception as e:
        logger.warning("Audit DB query failed: %s", e)

    # Get user departments from user store
    try:
        from auth.user_store import get_user_store
        user_store = await get_user_store()
        user_map = {u.username: u.department or "미배정" for u in await user_store.list_users()}
    except Exception:
        user_map = {}

    # Aggregate by user
    user_agg = {}
    for item in users_data:
        uname = item["username"]
        if uname not in user_agg:
            user_agg[uname] = {"username": uname, "query_count": 0, "department": user_map.get(uname, "")}
        user_agg[uname]["query_count"] += item["count"]

    # Aggregate by department
    for u in user_agg.values():
        dept = u.get("department", "미배정") or "미배정"
        dept_counter[dept] += u["query_count"]

    users = sorted(user_agg.values(), key=lambda x: x["query_count"], reverse=True)
    departments = [{"department": k, "query_count": v} for k, v in dept_counter.most_common()]

    return UserUsageStats(users=users, departments=departments)


@router.get("/export")
async def export_stats_excel(
    current_user: Annotated[User, Depends(require_role([Role.ADMIN, Role.MANAGER]))],
    period: Literal["day", "week", "month"] = "month",
):
    """통계 데이터 Excel 내보내기."""
    import openpyxl

    wb = openpyxl.Workbook()

    # Sheet 1: Usage
    ws1 = wb.active
    ws1.title = "사용량 통계"
    ws1.append(["기간", "총 세션", "총 질의", "고유 사용자", "평균 질의/세션"])

    usage = await get_usage_stats(current_user, period=period)
    ws1.append([usage.period, usage.total_sessions, usage.total_queries, usage.unique_users, usage.avg_queries_per_session])

    ws1.append([])
    ws1.append(["일자", "질의 수"])
    for d in usage.daily_breakdown:
        ws1.append([d["date"], d["queries"]])

    # Sheet 2: Keywords
    ws2 = wb.create_sheet("키워드 통계")
    ws2.append(["키워드", "횟수"])
    keywords = await get_keyword_stats(current_user, period=period, top_n=50)
    for kw in keywords.keywords:
        ws2.append([kw["keyword"], kw["count"]])

    # Sheet 3: User Usage (admin only)
    if current_user.role == Role.ADMIN:
        ws3 = wb.create_sheet("사용자별 통계")
        ws3.append(["사용자", "부서", "활동 수"])
        user_stats = await get_user_usage_stats(current_user, period=period)
        for u in user_stats.users:
            ws3.append([u["username"], u.get("department", ""), u["query_count"]])

        ws4 = wb.create_sheet("부서별 통계")
        ws4.append(["부서", "활동 수"])
        for d in user_stats.departments:
            ws4.append([d["department"], d["query_count"]])

    # Save to bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"flux-rag-stats-{period}-{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
